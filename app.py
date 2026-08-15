"""
MAS Exchange Rate - 30-Day Average Calculator (local web app)
----------------------------------------------------------------
Run this, then open the link it prints in your browser. 

This runs a tiny local web server on your own computer. The page in
your browser talks to this local server, and THIS SERVER (not your
browser) talks to the MAS API. That sidesteps the browser CORS block
you hit when the HTML file called MAS directly.

Needs only the Python standard library - no pip install required.

HOW TO RUN
----------
    python app.py

Then open the printed link (usually http://localhost:8765) in your
browser. Paste your MAS keyid into the page itself, same as before.
"""

import json
import os
import io
import webbrowser
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs, quote
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

PORT = int(os.environ.get("PORT", 8765))

MAS_BASE_URL = (
    "https://eservices.mas.gov.sg/apimg-gw/server/"
    "monthly_statistical_bulletin_non610ora/"
    "exchange_rates_end_of_period_daily/views/"
    "exchange_rates_end_of_period_daily"
)


def extract_rows(data):
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("result", "results", "data", "records", "elements", "rows"):
            if isinstance(data.get(key), list):
                return data[key]
        return [data]
    return []


def fetch_from_mas(date_str, field_name, keyid):
    """Call MAS server-side. Returns dict describing what happened."""
    url = f"{MAS_BASE_URL}?end_of_day={quote(date_str)}"
    req = Request(url, headers={"keyid": keyid})
    try:
        with urlopen(req, timeout=15) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
            status = resp.status
    except HTTPError as e:
        raw = e.read().decode("utf-8", errors="replace")
        return {"ok": False, "status": e.code, "raw": raw, "value": None}
    except URLError as e:
        return {"ok": False, "status": 0, "raw": f"Could not reach MAS: {e}", "value": None}

    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return {"ok": True, "status": status, "raw": raw, "value": None}

    rows = extract_rows(parsed)
    value = None
    if rows:
        row = rows[0]
        v = row.get(field_name) if isinstance(row, dict) else None
        if v not in (None, ""):
            try:
                value = float(v)
            except (TypeError, ValueError):
                value = None

    return {"ok": True, "status": status, "raw": raw, "value": value}


def build_excel(payload):
    currency = payload.get("currency", "")
    reply_date_str = payload.get("replyDateStr", "")
    window_start_str = payload.get("startStr", "")
    window_end_str = payload.get("endStr", "")
    display_rows = payload.get("displayBreakdown", [])  # list of [date, rate], full month context
    avg = payload.get("avg", None)

    wb = Workbook()
    ws = wb.active
    ws.title = "FX Average"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws.append(["MAS Exchange Rates - 30-Day Average"])
    ws["A1"].font = title_font
    ws.append([f"Vendor reply date: {reply_date_str}"])
    ws.append([f"Currency: {currency}"])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(["Date", "Rate"])
    ws.cell(row=header_row, column=1).font = bold
    ws.cell(row=header_row, column=2).font = bold

    for date_str, rate in display_rows:
        row_num = ws.max_row + 1
        ws.append([date_str, float(rate)])
        if window_start_str <= date_str <= window_end_str:
            ws.cell(row=row_num, column=1).fill = yellow_fill
            ws.cell(row=row_num, column=2).fill = yellow_fill

    ws.append([])
    ws.append([f"Average rate used is from {window_start_str} to {window_end_str}"])
    ws.cell(row=ws.max_row, column=1).font = bold

    if avg is not None:
        ws.append([f"Average rate: {round(float(avg), 4)}"])
        ws.cell(row=ws.max_row, column=1).font = bold

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


INDEX_HTML = None  # filled in at startup by reading index.html


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        pass  # keep console quiet

    def do_GET(self):
        parsed = urlparse(self.path)

        if parsed.path == "/" or parsed.path == "/index.html":
            body = INDEX_HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if parsed.path == "/api/rate":
            qs = parse_qs(parsed.query)
            date_str = qs.get("date", [""])[0]
            field = qs.get("field", [""])[0]
            keyid = qs.get("keyid", [""])[0]

            # basic validation of date format
            try:
                datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                self._send_json({"ok": False, "status": 0, "raw": "Bad date", "value": None}, 400)
                return

            if not keyid or not field:
                self._send_json({"ok": False, "status": 0, "raw": "Missing keyid or field", "value": None}, 400)
                return

            result = fetch_from_mas(date_str, field, keyid)
            self._send_json(result, 200)
            return

        self.send_response(404)
        self.end_headers()

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/export":
            length = int(self.headers.get("Content-Length", 0))
            raw_body = self.rfile.read(length)
            try:
                payload = json.loads(raw_body.decode("utf-8"))
            except json.JSONDecodeError:
                self.send_response(400)
                self.end_headers()
                return

            try:
                xlsx_bytes = build_excel(payload)
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "text/plain")
                self.end_headers()
                self.wfile.write(str(e).encode("utf-8"))
                return

            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="fx_average.xlsx"')
            self.send_header("Content-Length", str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)
            return

        self.send_response(404)
        self.end_headers()

    def _send_json(self, obj, status):
        body = json.dumps(obj).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


if __name__ == "__main__":
    with open("index.html", "r", encoding="utf-8") as f:
        INDEX_HTML = f.read()

    server = ThreadingHTTPServer(("0.0.0.0", PORT), Handler)
    url = f"http://localhost:{PORT}"
    print(f"Running on port {PORT}  (press Ctrl+C to stop)")
    if os.environ.get("PORT") is None:
        # only auto-open a browser tab when running on your own machine
        try:
            webbrowser.open(url)
        except Exception:
            pass
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")
